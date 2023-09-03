from datetime import datetime
from openpyxl.styles import Font, PatternFill
import csv
import openpyxl
import os
import psycopg2
import codecs
import re
import logging
import sys
import platform

# Local
PG_DB = "tally"
PG_HOST = "127.0.0.1"
PG_PORT = 5432
PG_USER = "postgres"
PG_PWD = "password"

if platform.system() == "Windows":
  OUTPUT_FONT_SIZE=12
else:
  OUTPUT_FONT_SIZE=20

def make_directories(filename):
  if not os.path.exists(os.path.dirname(filename)):
    os.makedirs(os.path.dirname(filename))

EXECUTION_TIME=datetime.now().strftime("%Y%m%d_%H%M%S")
LOG_DIR = os.path.join(os.path.dirname(__file__), 'logs')
LOG_FILE = LOG_DIR + "/" + EXECUTION_TIME + "_" + "gstvalidation.log"
RESULT_EXCEL_FILE = f"output{os.sep}validation_result_{EXECUTION_TIME}.xlsx"

make_directories(RESULT_EXCEL_FILE)
make_directories(LOG_FILE)

logging.basicConfig(
    level=logging.INFO,
    format='%(asctime)-15s %(levelname)s %(message)s',
    handlers=[
        logging.FileHandler(LOG_FILE),
        logging.StreamHandler(sys.stdout)
    ]
)

def process_portal_csv(filepath):
  logging.info(f"Processing file: {filepath}")
  with open(filepath, mode ='r') as file:
    # reading the CSV file
    csvFile = csv.reader(file)
    values = []

    # skipping first 3 lines: 2 lines + 3rd header line
    for i in range(3): next(csvFile)

    # displaying the contents of the CSV file
    for lines in csvFile:
      if len(lines) != 0:
        values.append(f"('{lines[0].strip()}','{lines[1].strip()}','{lines[3].strip()}','{lines[4].strip()}')")

    if len(values) !=0:
      return f"INSERT INTO gst_invoices(gstin,number,date,value) VALUES {','.join(values)}"

    return None

def process_tally_excel(filepath):
  logging.info(f"Processing file: {filepath}")
  dataframe = openpyxl.load_workbook(filepath)

  # Define variable to read sheet
  xlsheet = dataframe.active

  headers = ["date,particulars,supplier,invoice_number,gstin,value,gross_total"]
  values = []
  # Iterate the loop to read the cell values
  for row in range(8, xlsheet.max_row):
    value = []
    for col in range(1,8):
      value.append(f"'{str(xlsheet.cell(row=row, column=col).value).strip()}'")
    values.append(f"({','.join(value)})")

  if len(values) !=0:
   return f"INSERT INTO tally_invoices(date,particulars,supplier,invoice_number,gstin,value,gross_total) VALUES {','.join(values)}"

  return None

def read_sql_file(file):
  with codecs.open(file, mode='r', encoding='utf-8', buffering=-1) as sql_file:
    return sql_file.read()

def process_directory_files(type,directory):
  # iterate over files in
  # that directory
  sqls = []

  for filename in os.listdir(directory):
    file = os.path.join(directory, filename)
    # checking if it is a file
    if os.path.isfile(file) and not filename.startswith("~"):
      if type == "PORTAL" and filename.endswith(".csv") :
        sqls.append(process_portal_csv(file))
      elif type == "TALLY" and filename.endswith(".xlsx"):
        sqls.append(process_tally_excel(file))
      elif filename.endswith(".sql"):
        sqls.append(read_sql_file(file))

  return sqls

def execute_postgres_sqls(list_of_sqls):
  try:
    conn = psycopg2.connect(database=PG_DB, user=PG_USER, password=PG_PWD, host=PG_HOST, port=PG_PORT)
    # conn = jaydebeapi.connect("org.postgresql.Driver",
    #                         "jdbc:postgresql://127.0.0.1:5432/returns",
    #                           ["root", "root"],
    #                         "postgresql-42.6.0.jar",)
    # logging.info("Database opened successfully")
    #  The except block lets you handle the error.

    #Creating a cursor object using the cursor() method
    cursor = conn.cursor()
    data = []
    colnames = []
    for sql in list_of_sqls:
      logging.info(f"Executing SQL: {sql}")
      cursor.execute(sql)
      conn.commit()
      if sql.startswith("SELECT"):
        data.append(cursor.fetchall())
        colnames = [desc[0] for desc in cursor.description]
    return {"colnames": colnames, "data": data}

  except Exception as error:
    logging.info(error)

  finally:
    if (conn):
      cursor.close()
      conn.close()
      logging.info("PostgreSQL connection is closed")

def isequal_invoice_numbers(gst_num,tally_num):
  is_eq = False
  if gst_num and tally_num:
    is_eq = (tally_num == gst_num) \
            or (tally_num.lstrip('0') == gst_num.lstrip('0')) \
            or (tally_num.split("/",1)[0].lstrip('0') == gst_num.split("/",1)[0].lstrip('0') ) \
            or (tally_num.split("/",1)[-1].lstrip('0') == gst_num.split("/",1)[-1].lstrip('0')) \
            or (tally_num.__contains__(gst_num)) \
            or tally_num.startswith(re.findall('[0-9]+', gst_num)[0])

  return is_eq

def store_result_in_excel(filename,result,title):
  logging.info(f"Storing data for {title} in excel file {filename}")
  data = result["data"]
  colnames = result["colnames"]
  if os.path.isfile(filename):
    wBook = openpyxl.load_workbook(filename)
    try:
      sheet = wBook[title]
    except Exception as error:
      sheet = wBook.create_sheet(title)
  else:
    wBook = openpyxl.Workbook()
    sheet = wBook.create_sheet(title)

  # sheet.append(["supplier","gstin","gst_date","tally_date", "gst_invoice_number","tally_invoice_number","tally_invoice_value","tally_gross_total","gst_invoice_value"])
  sheet.append(colnames)
  for row in data[0]:
    sheet.append(row)

  header_font = Font(color='00FFFFFF',bold=True,size=OUTPUT_FONT_SIZE)
  header_fill = PatternFill("solid", fgColor="003366FF")

  # Enumerate the cells in the first row
  for cell in sheet["1:1"]:
    cell.font = header_font
    cell.fill = header_fill

  if 'Sheet' in wBook.sheetnames:
    wBook.remove(wBook['Sheet'])

  for row in sheet.iter_cols(min_row=2, min_col=1, max_row=sheet.max_row, max_col=sheet.max_column):
    for cell in row:
      cell.font = Font(size=OUTPUT_FONT_SIZE)

  if title == "All Invoices":
    for i in range(2, sheet.max_row):
      if not isequal_invoice_numbers(sheet['E{}'.format(i)].value,sheet['F{}'.format(i)].value):
        sheet['E{}'.format(i)].font = Font(size=OUTPUT_FONT_SIZE,color='00FF0000',bold=True)
        sheet['F{}'.format(i)].font = Font(size=OUTPUT_FONT_SIZE,color='00FF0000',bold=True)

  wBook.save(filename)

if __name__ == "__main__":
  logging.info("Fetching all SQLs")
  list_of_sqls = [] + \
    process_directory_files("DDLS",f"queries{os.sep}postgres{os.sep}ddls" ) + \
    process_directory_files("PORTAL",f"exports{os.sep}portal" ) + \
    process_directory_files("TALLY",f"exports{os.sep}tally" ) + \
    process_directory_files("VALIDATION",f"queries{os.sep}postgres{os.sep}base_tables" ) + \
    process_directory_files("STATS",f"queries{os.sep}postgres{os.sep}stats" )

  logging.info(f"List of All SQLs to be executed:- {os.linesep}{os.linesep.join(list_of_sqls)}")
  data = execute_postgres_sqls(list_of_sqls)

  logging.info("Fetching all Non Matched records")
  result = execute_postgres_sqls(["SELECT * FROM invoices_not_matched"])
  store_result_in_excel(RESULT_EXCEL_FILE,result,"Not Matched Invoices")

  logging.info("Fetching Not matched invoices stats")
  result = execute_postgres_sqls(["SELECT * FROM not_matched_invoices_stats"])
  store_result_in_excel(RESULT_EXCEL_FILE,result,"Not Matched Invoices Stats")

  logging.info("Fetching all Matched records")
  result = execute_postgres_sqls(["SELECT * FROM invoices_matched"])
  store_result_in_excel(RESULT_EXCEL_FILE,result,"Matched Invoices")

  logging.info("Fetching all invoices records")
  result = execute_postgres_sqls(["SELECT * FROM all_invoices"])
  store_result_in_excel(RESULT_EXCEL_FILE,result,"All Invoices")

  logging.info("Fetching all invoices stats")
  result = execute_postgres_sqls(["SELECT * FROM all_invoices_stats"])
  store_result_in_excel(RESULT_EXCEL_FILE,result,"All Invoices Stats")

